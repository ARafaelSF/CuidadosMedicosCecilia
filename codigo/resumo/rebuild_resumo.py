# -*- coding: utf-8 -*-
"""Auditoria + reconstrução do Resumo Exames.xlsm."""
from __future__ import annotations

import re
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\arafa\OneDrive\Documentos\Cecília\_Organizado")
EXAMES = BASE / "Exames"
RESUMO = EXAMES / "Resumo Exames.xlsm"
OUT_XLSX = Path(r"C:\Users\arafa\AppData\Local\Temp\ResumoExames.xlsx")

TIPO_COLORS = {
    "Sangue": "D6EAF8",
    "Imagem": "D5F5E3",
    "Audiologia": "FCF3CF",
    "EEG": "E8DAEF",
    "Pezinho": "FADBD8",
    "Urina": "D6EAF8",
    "Suor": "FAE5D3",
}

# Laudos Hermes Pardini (conferidos) vs anteriores (pendente)
# Datas de sangue não-Hermes tipicamente 2020–2022 até antes do padrão Hermes.
def conferido_default(tipo: str, d: date | None, arquivo: str) -> str:
    return "Sim"


def parse_pdf_name(name: str):
    stem = name[:-4] if name.lower().endswith(".pdf") else name
    m = re.match(r"^(.+?) - (\d{4}-\d{2}-\d{2}) - (.+)$", stem)
    if not m:
        return None
    tipo, ds, rest = m.group(1), m.group(2), m.group(3)
    d = datetime.strptime(ds, "%Y-%m-%d").date()
    return tipo, d, rest, stem


def norm(s: str) -> str:
    s = (s or "").casefold()
    rep = str.maketrans(
        "áàâãäéèêëíìîïóòôõöúùûüçñ",
        "aaaaaeeeeiiiiooooouuuucn",
    )
    s = s.translate(rep)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def first_names(prof: str) -> str:
    # "Dra. Júlia" / "Aline Cornelio" -> primeiras palavras úteis
    p = re.sub(r"^(dr\.?a?|dra\.?)\s*", "", (prof or "").casefold()).strip()
    parts = [x for x in re.split(r"\s+", p) if x and x not in {"de", "da", "do", "dos", "das"}]
    return " ".join(parts[:2]) if parts else ""


def score_match(row_tipo, row_date, row_prof, row_desc, pdf_tipo, pdf_date, pdf_rest) -> int:
    sc = 0
    if norm(row_tipo) == norm(pdf_tipo):
        sc += 50
    else:
        return -1
    if row_date and pdf_date and row_date == pdf_date:
        sc += 40
    else:
        return -1
    rn = norm(first_names(row_prof))
    pn = norm(pdf_rest)
    dn = norm(row_desc)
    if not rn:
        if "sem solicitante" in pn:
            sc += 20
    else:
        tokens = rn.split()
        hit = sum(1 for t in tokens if t in pn)
        if hit:
            sc += 10 * hit
        elif "sem solicitante" in pn and (not row_prof or row_prof.strip() in {"—", "-", "–", "?"}):
            sc += 15

    # Desambiguação no mesmo dia pelo sufixo/descrição
    keywords = [
        ("eoapd", ["eoapd", "eoa pd"]),
        ("eoat", ["eoat", "eoa t"]),
        ("audio", ["audiometria", "imitancio", "imitancia"]),
        ("peate", ["peate"]),
        ("fator xiii", ["fator xiii", "fator xiii"]),
        ("grupo", ["grupo sanguineo", "grupo rh", "fator rh"]),
        ("culturas", ["cultura", "hemocultura", "gram"]),
        ("punho", ["idade ossea", "punho", "mao"]),
        ("tc cranio", ["tc de cranio", "tc cranio"]),
    ]
    for key, desc_keys in keywords:
        in_pdf = key in pn or any(k in pn for k in desc_keys)
        in_desc = any(k in dn for k in desc_keys) or key in dn
        if in_pdf and in_desc:
            sc += 25
        elif in_pdf and not in_desc:
            sc -= 5
    return sc


def load_current_rows():
    wb = openpyxl.load_workbook(RESUMO, keep_vba=True, data_only=True)
    try:
        ws = wb["Resumo"]
        rows = []
        for r in range(3, ws.max_row + 1):
            d = ws.cell(r, 1).value
            tipo = ws.cell(r, 2).value
            prof = ws.cell(r, 3).value
            desc = ws.cell(r, 4).value
            if not tipo:
                continue
            if isinstance(d, datetime):
                d = d.date()
            elif not isinstance(d, date):
                continue
            prof_s = str(prof or "")
            # Correção conhecida: teste do suor foi pedido pelo Guilherme Rache
            if str(tipo) == "Suor" and d == date(2024, 10, 18):
                prof_s = "Dr. Guilherme Rache"
            rows.append({"data": d, "tipo": str(tipo or ""), "prof": prof_s, "desc": str(desc or "")})
        return rows
    finally:
        wb.close()


def main():
    rows = load_current_rows()
    pdfs = sorted(
        [p for p in EXAMES.glob("*.pdf") if p.is_file() and not p.name.startswith("_")],
        key=lambda p: p.name.casefold(),
    )
    parsed = []
    for p in pdfs:
        info = parse_pdf_name(p.name)
        if info:
            tipo, d, rest, stem = info
            parsed.append({"path": p, "tipo": tipo, "data": d, "rest": rest, "stem": stem, "name": p.name})
        else:
            print("UNPARSED", p.name)

    used = set()
    matched = [None] * len(rows)
    pairs = []
    for i, row in enumerate(rows):
        for j, pdf in enumerate(parsed):
            sc = score_match(
                row["tipo"], row["data"], row["prof"], row["desc"],
                pdf["tipo"], pdf["data"], pdf["rest"],
            )
            if sc >= 90:
                pairs.append((sc, i, j))
    pairs.sort(key=lambda x: (-x[0], x[1], x[2]))
    for sc, i, j in pairs:
        if matched[i] is not None or j in used:
            continue
        used.add(j)
        pdf = parsed[j]
        row = rows[i]
        conf = conferido_default(row["tipo"], row["data"], pdf["name"])
        matched[i] = {**row, "arquivo": pdf["stem"], "conferido": conf, "stem": pdf["stem"], "pdf": pdf["name"]}
        print(f"OK {row['data']} {row['tipo'][:12]:12} -> {pdf['name'][:70]} (sc={sc})")

    for i, row in enumerate(rows):
        if matched[i] is not None:
            continue
        # fallback tipo+data
        best = None
        for j, pdf in enumerate(parsed):
            if j in used:
                continue
            if norm(pdf["tipo"]) == norm(row["tipo"]) and pdf["data"] == row["data"]:
                best = j
                break
        if best is None:
            print("NO_MATCH", row)
            matched[i] = {**row, "arquivo": "", "conferido": "Pendente", "stem": "", "pdf": ""}
        else:
            used.add(best)
            pdf = parsed[best]
            conf = conferido_default(row["tipo"], row["data"], pdf["name"])
            matched[i] = {**row, "arquivo": pdf["stem"], "conferido": conf, "stem": pdf["stem"], "pdf": pdf["name"]}
            print(f"FALLBACK {row['data']} {row['tipo'][:12]:12} -> {pdf['name'][:70]}")

    unused = [parsed[j]["name"] for j in range(len(parsed)) if j not in used]
    print("ROWS", len(rows), "PDFS", len(pdfs), "UNUSED", len(unused))
    for u in unused:
        print(" UNUSED", u)

    matched = list(matched)

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    title = "Cecília Maria Albergaria Silva — Resumo de exames"
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Data", "Tipo", "Profissional", "Descrição", "Arquivo"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    body_font = Font(name="Calibri", size=10, color="333333")
    link_font = Font(name="Calibri", size=9, color="0563C1", underline="single")
    # Sem fill nas células de dados (exceto Tipo): o zebrado vem da Tabela no Excel.
    no_fill = PatternFill(fill_type=None)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[2].height = 20

    # Sort by Tipo then Data (current sort state was Tipo)
    matched.sort(key=lambda x: (x["tipo"], x["data"] or date.min, x["arquivo"]))

    for i, row in enumerate(matched):
        r = i + 3
        tipo_fill = PatternFill("solid", fgColor=TIPO_COLORS.get(row["tipo"], "FFFFFF"))

        ws.cell(r, 1, row["data"]).number_format = "DD/MM/YYYY"
        ws.cell(r, 2, row["tipo"])
        ws.cell(r, 3, row["prof"])
        ws.cell(r, 4, row["desc"])
        ws.cell(r, 5, row["arquivo"])

        for col in range(1, 6):
            cell = ws.cell(r, col)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))
            if col == 2:
                cell.fill = tipo_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = no_fill
                if col == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # hyperlink to PDF
        if row.get("pdf"):
            cell = ws.cell(r, 5)
            cell.hyperlink = row["pdf"]  # relative to workbook folder
            cell.font = link_font
            cell.fill = no_fill

        # row height estimate
        desc_len = len(row["desc"] or "")
        ws.row_dimensions[r].height = min(120, max(18, 14 + (desc_len // 70) * 12))

    last = 2 + len(matched)

    # NÃO criar Table via openpyxl (costuma gerar xlsm/xlsx que o Excel recusa abrir).
    # A tabela nativa + zebrado são aplicados no rebuild_xlsm.ps1 via COM.

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 60
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:E{last}"

    # Print
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    note = ws.cell(last + 2, 1, "Arquivo = PDF na pasta Exames (sem .pdf).")
    note.font = Font(name="Calibri", size=9, italic=True, color="666666")
    ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=5)

    wb.save(OUT_XLSX)
    wb.close()
    # Cópia só no TEMP — o ativo é o .xlsm (macros). Evita dois arquivos em Exames\.
    print("SAVED", OUT_XLSX)
    print("ROWS", len(matched))


if __name__ == "__main__":
    main()
