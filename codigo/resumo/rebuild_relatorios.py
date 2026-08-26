# -*- coding: utf-8 -*-
"""Reconstrução do Resumo Relatórios.xlsm (mesmo padrão do Resumo Exames)."""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = Path(r"C:\Users\arafa\OneDrive\Documentos\Cecília\_Organizado")
REL = BASE / "Relatórios"
RESUMO_SRC = REL / "Resumo Relatórios.xlsx"
RESUMO_XLSM = REL / "Resumo Relatórios.xlsm"
OUT_XLSX = Path(r"C:\Users\arafa\AppData\Local\Temp\ResumoRelatorios.xlsx")

TIPO_COLORS = {
    "Alta": "FADBD8",
    "Endocrino": "D6EAF8",
    "Escola": "FCF3CF",
    "Fisioterapia": "D5F5E3",
    "Fonoaudiologia": "E8DAEF",
    "Funcional": "D5F5E3",
    "Genética": "D6EAF8",
    "Natação": "D4E6F1",
    "Neuro": "E8DAEF",
    "Oftalmologia": "FCF3CF",
    "Pedagogia": "FAE5D3",
    "Psicologia": "E8DAEF",
    "TO": "D5F5E3",
    "TO Visão": "D5F5E3",
}


def parse_pdf_name(name: str):
    stem = name[:-4] if name.lower().endswith(".pdf") else name
    m = re.match(r"^(.+?) - (\d{4}-\d{2}-\d{2}) - (.+)$", stem)
    if not m:
        return None
    tipo, ds, rest = m.group(1), m.group(2), m.group(3)
    d = datetime.strptime(ds, "%Y-%m-%d").date()
    return tipo, d, rest, stem


def norm(s: str) -> str:
    s = (s or "").casefold()
    rep = str.maketrans(
        "áàâãäéèêëíìîïóòôõöúùûüçñ",
        "aaaaaeeeeiiiiooooouuuucn",
    )
    s = s.translate(rep)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def first_tokens(prof: str) -> str:
    p = re.sub(r"^(dr\.?a?|dra\.?)\s*", "", (prof or "").casefold()).strip()
    parts = [x for x in re.split(r"\s+", p) if x and x not in {"de", "da", "do", "dos", "das", "e"}]
    return " ".join(parts[:2]) if parts else ""


def score_match(row_tipo, row_date, row_prof, row_desc, pdf_tipo, pdf_date, pdf_rest) -> int:
    sc = 0
    if norm(row_tipo) != norm(pdf_tipo):
        return -1
    sc += 50
    if not row_date or row_date != pdf_date:
        return -1
    sc += 40
    rn = norm(first_tokens(row_prof))
    pn = norm(pdf_rest)
    dn = norm(row_desc)
    if rn:
        hit = sum(1 for t in rn.split() if t in pn)
        if hit:
            sc += 10 * hit
    keywords = [
        ("enfermagem", ["enfermagem"]),
        ("exoma", ["exoma"]),
        ("casu", ["casu"]),
        ("neuro", ["neuro", "neurolog"]),
        ("marlon", ["marlon"]),
        ("bayley", ["bayley"]),
        ("pauta", ["pauta"]),
        ("periodo", ["periodo", "período"]),
        ("semestre", ["semestre"]),
        ("trimestre", ["trimestre"]),
        ("emei", ["emei"]),
        ("avif", ["avif", "visao funcional", "visão funcional"]),
    ]
    for key, desc_keys in keywords:
        in_pdf = key in pn or any(k in pn for k in desc_keys)
        in_desc = key in dn or any(k in dn for k in desc_keys)
        if in_pdf and in_desc:
            sc += 25
        elif in_pdf and not in_desc:
            sc -= 3
    return sc


def _header_map(ws) -> dict[str, int]:
    out = {}
    for c in range(1, 12):
        v = ws.cell(2, c).value
        if v:
            out[str(v).strip().casefold()] = c
    return out


def load_current_rows():
    src = RESUMO_XLSM if RESUMO_XLSM.exists() else RESUMO_SRC
    wb = openpyxl.load_workbook(src, data_only=True)
    try:
        ws = wb["Resumo"] if "Resumo" in wb.sheetnames else wb.active
        headers = _header_map(ws)
        col_conf = headers.get("conferido")
        col_imp = headers.get("impresso")
        rows = []
        for r in range(3, ws.max_row + 1):
            d = ws.cell(r, 1).value
            tipo = ws.cell(r, 2).value
            prof = ws.cell(r, 3).value
            desc = ws.cell(r, 4).value
            if not tipo:
                continue
            if isinstance(d, datetime):
                d = d.date()
            elif not isinstance(d, date):
                # Ebook / material sem data — fora do escopo de Relatórios
                print("SKIP_NO_DATE", tipo, desc)
                continue
            desc_s = str(desc or "")
            if "ebook" in desc_s.casefold() or "estimula beb" in desc_s.casefold():
                print("SKIP_MATERIAL", desc_s[:60])
                continue
            conf = str(ws.cell(r, col_conf).value or "Pendente").strip() if col_conf else "Pendente"
            imp = str(ws.cell(r, col_imp).value or "Pendente").strip() if col_imp else "Pendente"
            if conf not in ("Sim", "Pendente", "Não"):
                conf = "Pendente"
            if imp not in ("Sim", "Pendente", "Não"):
                imp = "Pendente"
            rows.append({
                "data": d,
                "tipo": str(tipo),
                "prof": str(prof or ""),
                "desc": desc_s,
                "conferido": conf,
                "impresso": imp,
            })
        return rows
    finally:
        wb.close()


def main():
    rows = load_current_rows()
    pdfs = sorted(
        [
            p for p in REL.glob("*.pdf")
            if p.is_file() and not p.name.startswith("_") and not p.name.casefold().startswith("imprimir")
        ],
        key=lambda p: p.name.casefold(),
    )
    parsed = []
    for p in pdfs:
        info = parse_pdf_name(p.name)
        if not info:
            print("UNPARSED", p.name)
            continue
        tipo, d, rest, stem = info
        parsed.append({"path": p, "tipo": tipo, "data": d, "rest": rest, "stem": stem, "name": p.name})

    matched = [None] * len(rows)
    used = set()
    pairs = []
    for i, row in enumerate(rows):
        for j, pdf in enumerate(parsed):
            sc = score_match(
                row["tipo"], row["data"], row["prof"], row["desc"],
                pdf["tipo"], pdf["data"], pdf["rest"],
            )
            if sc >= 90:
                pairs.append((sc, i, j))
    pairs.sort(key=lambda x: (-x[0], x[1], x[2]))
    for sc, i, j in pairs:
        if matched[i] is not None or j in used:
            continue
        used.add(j)
        pdf = parsed[j]
        row = rows[i]
        matched[i] = {
            **row,
            "arquivo": pdf["stem"],
            "pdf": pdf["name"],
            "conferido": row.get("conferido", "Pendente"),
            "impresso": row.get("impresso", "Pendente"),
        }
        print(f"OK {row['data']} {row['tipo'][:14]:14} -> {pdf['name'][:65]} (sc={sc})")

    for i, row in enumerate(rows):
        if matched[i] is not None:
            continue
        best = None
        for j, pdf in enumerate(parsed):
            if j in used:
                continue
            if norm(pdf["tipo"]) == norm(row["tipo"]) and pdf["data"] == row["data"]:
                best = j
                break
        if best is None:
            print("NO_MATCH", row)
            matched[i] = {
                **row,
                "arquivo": "",
                "pdf": "",
                "conferido": row.get("conferido", "Pendente"),
                "impresso": row.get("impresso", "Pendente"),
            }
        else:
            used.add(best)
            pdf = parsed[best]
            matched[i] = {
                **row,
                "arquivo": pdf["stem"],
                "pdf": pdf["name"],
                "conferido": row.get("conferido", "Pendente"),
                "impresso": row.get("impresso", "Pendente"),
            }
            print(f"FALLBACK {row['data']} {row['tipo'][:14]:14} -> {pdf['name'][:65]}")

    unused = [parsed[j] for j in range(len(parsed)) if j not in used]
    print("ROWS", len(rows), "PDFS", len(pdfs), "UNUSED", len(unused))
    matched = list(matched)

    # PDFs novos (ainda sem linha no resumo)
    DESC_NOVOS = {
        "Fisioterapia - 2026-06-01 - Silvia Figueiredo": "Alta definitiva (reavaliação 27/05/2026)",
        "Fisioterapia - 2026-06-01 - Silvia Figueiredo - Marlon": "Relatório ao educador físico Marlon (alta definitiva)",
        "Escola - 2023-12-31 - EMEI Alaíde Lisboa": "Relatório individual 2º semestre/2023",
        "Fisioterapia - 2023-01-27 - Silvia Figueiredo": "Evolução motora (Estimula)",
        "Fisioterapia - 2022-03-20 - Silvia Figueiredo - Neuro": "Evolução motora (à neurologista)",
        "Fisioterapia - 2022-10-28 - Silvia Figueiredo": "Relatório de sessões (set–out/2022)",
        "Fonoaudiologia - 2024-03-12 - Viviane Cardoso Sampaio": "Relatório de acompanhamento / continuidade",
        "Fonoaudiologia - 2024-04-03 - Fernanda Alvarenga de Castro": "Acompanhamento (Travessia; ABFW)",
        "TO - 2024-04-03 - Manon Faria": "Acompanhamento TO / PEDI — alta semanal (Travessia)",
    }
    PROF_NOVOS = {
        "Escola - 2023-12-31 - EMEI Alaíde Lisboa": "EMEI Alaíde Lisboa",
        "Fonoaudiologia - 2024-03-12 - Viviane Cardoso Sampaio": "Viviane Sampaio",
        "Fonoaudiologia - 2024-04-03 - Fernanda Alvarenga de Castro": "Fernanda Alvarenga",
        "TO - 2024-04-03 - Manon Faria": "Manon Faria",
        "Fisioterapia - 2023-01-27 - Silvia Figueiredo": "Silvia Figueiredo",
        "Fisioterapia - 2022-03-20 - Silvia Figueiredo - Neuro": "Silvia Figueiredo",
        "Fisioterapia - 2022-10-28 - Silvia Figueiredo": "Silvia Figueiredo",
    }
    for pdf in unused:
        stem = pdf["stem"]
        desc = DESC_NOVOS.get(stem)
        if not desc:
            if " - " in pdf["rest"]:
                suf = pdf["rest"].split(" - ", 1)[1]
                desc = f"Relatório ({suf})"
            else:
                desc = "Relatório"
        prof_full = pdf["rest"].split(" - ", 1)[0]
        if stem in PROF_NOVOS:
            prof = PROF_NOVOS[stem]
        elif "Silvia Figueiredo" in prof_full:
            prof = "Silvia Figueiredo"
        else:
            prof = prof_full
        matched.append({
            "data": pdf["data"],
            "tipo": pdf["tipo"],
            "prof": prof,
            "desc": desc,
            "arquivo": stem,
            "pdf": pdf["name"],
            "conferido": "Pendente",
            "impresso": "Pendente",
        })
        print(f"NOVO {pdf['data']} {pdf['tipo'][:14]:14} -> {pdf['name'][:65]}")

    matched.sort(key=lambda x: (x["tipo"], x["data"] or date.min, x["arquivo"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Cecília Maria Albergaria Silva — Resumo de relatórios"
    c.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Data", "Tipo", "Profissional", "Descrição", "Arquivo", "Conferido", "Impresso"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    body_font = Font(name="Calibri", size=10, color="333333")
    link_font = Font(name="Calibri", size=9, color="0563C1", underline="single")
    pend_font = Font(name="Calibri", size=10, color="B9770E", bold=True)
    ok_font = Font(name="Calibri", size=10, color="1E8449", bold=True)
    no_fill = PatternFill(fill_type=None)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[2].height = 20

    for i, row in enumerate(matched):
        r = i + 3
        tipo_fill = PatternFill("solid", fgColor=TIPO_COLORS.get(row["tipo"], "FFFFFF"))
        ws.cell(r, 1, row["data"]).number_format = "DD/MM/YYYY"
        ws.cell(r, 2, row["tipo"])
        ws.cell(r, 3, row["prof"])
        ws.cell(r, 4, row["desc"])
        ws.cell(r, 5, row["arquivo"])
        ws.cell(r, 6, row["conferido"])
        ws.cell(r, 7, row["impresso"])
        for col in range(1, 8):
            cell = ws.cell(r, col)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))
            if col == 2:
                cell.fill = tipo_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = no_fill
                if col in (1, 6, 7):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if col in (6, 7):
                    val = str(cell.value or "")
                    if val == "Sim":
                        cell.font = ok_font
                    else:
                        cell.font = pend_font
        if row.get("pdf"):
            cell = ws.cell(r, 5)
            cell.hyperlink = row["pdf"]
            cell.font = link_font
            cell.fill = no_fill
        desc_len = len(row["desc"] or "")
        ws.row_dimensions[r].height = min(110, max(18, 14 + (desc_len // 60) * 12))

    last = 2 + len(matched)

    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Sim,Pendente,Não"', allow_blank=True)
    dv.errorTitle = "Status"
    dv.error = "Escolha Sim, Pendente ou Não"
    ws.add_data_validation(dv)
    dv.add(f"F3:G{last}")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{last}"

    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    note = ws.cell(
        last + 2, 1,
        "Arquivo = PDF na pasta Relatórios (sem .pdf). Conferido / Impresso: marque Sim depois de validar ou imprimir.",
    )
    note.font = Font(name="Calibri", size=9, italic=True, color="666666")
    ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=7)

    wb.save(OUT_XLSX)
    wb.close()
    print("SAVED", OUT_XLSX)
    print(
        "ROWS", len(matched),
        "CONF_PEND", sum(1 for m in matched if m["conferido"] == "Pendente"),
        "IMP_PEND", sum(1 for m in matched if m["impresso"] == "Pendente"),
    )


if __name__ == "__main__":
    main()
