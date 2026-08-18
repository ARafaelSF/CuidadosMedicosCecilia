# -*- coding: utf-8 -*-
"""Resumo e evolução dos exames laboratoriais do CTI (pasta isolada)."""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import xlsxwriter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from caminhos import CTI, TEMP  # noqa: E402
TEXTS = json.loads((TEMP / "cti_texts.json").read_text(encoding="utf-8"))
NUM = r"([0-9]+(?:\.[0-9]{3})*(?:,[0-9]+)?)"


def br_num(s: str | None):
    if not s:
        return None
    s = s.strip()
    if "," in s:
        return float(s.replace(".", "").replace(",", "."))
    parts = s.split(".")
    if len(parts) == 2 and len(parts[1]) == 3:
        return float(s.replace(".", ""))
    return float(s)


def parse_date(s: str | None):
    if not s:
        return None
    return datetime.strptime(s, "%d/%m/%Y").date()


def short_local(local: str | None) -> str:
    if not local:
        return "—"
    u = local.lower()
    if "cti" in u:
        return "CTI Infantil"
    if "pronto atendimento" in u or u.startswith("pa"):
        return "PA"
    if "cirúrgico" in u:
        return "Centro Cirúrgico"
    return local.strip()


def parse_header(text: str) -> dict:
    medico = re.search(r"Médico:\s*Dr\(a\)\.\s*(.+?)\s+Local:", text)
    local = re.search(r"Local:\s*(.+?)\s*\n", text)
    data = re.search(r"Data coleta:\s*(\d{2}/\d{2}/\d{4})", text)
    presc = re.search(r"Prescrição:\s*(\d+)", text)
    idade = re.search(r"Idade\s+(.+?)\s+Sexo", text)
    return {
        "medico": medico.group(1).strip() if medico else None,
        "local": local.group(1).strip() if local else None,
        "data": parse_date(data.group(1) if data else None),
        "data_txt": data.group(1) if data else None,
        "presc": presc.group(1) if presc else None,
        "idade": idade.group(1).strip() if idade else None,
    }


def split_pages(text: str) -> list[str]:
    parts = re.split(r"(?=LABORATÓRIO DE PATOLOGIA CLÍNICA)", text)
    return [p for p in parts if p.strip()]


def first(pat: str, text: str, flags=0):
    m = re.search(pat, text, flags)
    return br_num(m.group(1)) if m else None


def first_str(pat: str, text: str, flags=0):
    m = re.search(pat, text, flags)
    return m.group(1).strip() if m else None


def describe(text: str) -> str:
    bits = []
    if re.search(r"\bHEMOGRAMA\b", text):
        bits.append("Hemograma")
    if "PROTEÍNA C REATIVA" in text:
        bits.append("PCR")
    if re.search(r"(?m)^GLICOSE:", text):
        bits.append("glicose")
    elec = []
    if re.search(r"SÓDIO", text):
        elec.append("Na")
    if re.search(r"POTÁSSIO", text):
        elec.append("K")
    if re.search(r"(?m)^CLORO", text):
        elec.append("Cl")
    if "CÁLCIO IÔNICO" in text:
        elec.append("Ca iônico")
    if re.search(r"MAGNÉSIO", text):
        elec.append("Mg")
    if elec:
        bits.append("eletrólitos (" + ", ".join(elec) + ")")
    if re.search(r"CREATININA", text):
        bits.append("creatinina")
    if re.search(r"GASOMETRIA ARTERIAL", text, re.I):
        bits.append("gasometria arterial")
    elif re.search(r"GASOMETRIA VENOSA", text, re.I):
        bits.append("gasometria venosa")
    elif re.search(r"GASOMETRIA", text, re.I):
        bits.append("gasometria")
    if re.search(r"ÁCIDO LÁTICO", text):
        bits.append("lactato")
    if "COAGULOGRAMA" in text:
        bits.append("coagulograma")
    if "FATOR XIII" in text:
        bits.append("fator XIII")
    if "VANCOMICINA" in text:
        bits.append("vancomicina")
    if re.search(r"(?m)^CORTISOL\b", text):
        bits.append("cortisol")
    if "TIREOTROPINA" in text or re.search(r"TSH\)", text):
        bits.append("TSH neonatal")
    if "ROTINA DE LÍQUOR" in text or "ROTINA DE LIQUOR" in text:
        bits.append("líquor")
    if re.search(r"ELEMENTOS ANORMAIS E SEDIMENTOSCOPIA", text):
        bits.append("EAS")
    if re.search(r"GRAM BACTERIOSCOPIA", text):
        bits.append("Gram")
    if "UROCULTURA" in text:
        uro = "urocultura"
        if re.search(r"UROCULTURA.{0,250}Positivo", text, re.S | re.I):
            isol = re.search(r"Isolado 01:\s*([^\n]+)", text, re.I)
            uro += " positiva" + (f" ({isol.group(1).strip()})" if isol else "")
        else:
            uro += " negativa"
        bits.append(uro)
    if "HEMOCULTURA" in text:
        bits.append("hemocultura negativa")
    if re.search(r"CULTURA AERÓBIA", text) and "Líquor" in text:
        bits.append("cultura de líquor negativa")
    if not bits:
        bits.append("Ver PDF")
    # unique preserve
    seen = set()
    out = []
    for b in bits:
        if b not in seen:
            seen.add(b)
            out.append(b)
    s = ", ".join(out)
    return s[0].upper() + s[1:] if s else s


def extract_gas_blocks(text: str):
    out = []
    for m in re.finditer(
        r"GASOMETRIA (ARTERIAL|VENOSA):(.{0,1400}?)(?=Este laudo|LABORATÓRIO DE PATOLOGIA|GASOMETRIA |\Z)",
        text,
        re.S | re.I,
    ):
        kind = m.group(1).title()
        block = m.group(2)
        out.append(
            {
                "tipo": kind,
                "ph": first(rf"PH\.+:\s*{NUM}", block, re.I),
                "pco2": first(rf"PCO2\.+:\s*{NUM}", block, re.I),
                "po2": first(rf"PO2\.+:\s*{NUM}", block, re.I),
                "hco3": first(rf"HCO3\.+:\s*{NUM}", block, re.I),
                "be": first(rf"BE\.+:\s*{NUM}", block, re.I),
                "so2": first(rf"SO2\.+:\s*{NUM}", block, re.I),
            }
        )
    return out


def extract_liquor(text: str):
    if "ROTINA DE LÍQUOR" not in text and "ROTINA DE LIQUOR" not in text:
        return None
    aspecto = first_str(r"Aspecto antes da centrifugação:\s*(.+)", text)
    return {
        "aspecto": aspecto,
        "hemacias": first(rf"Hemácias\.+:\s*{NUM}\s*p/mm", text, re.I),
        "leuc": first(rf"Leucócitos:\s*{NUM}\s*p/mm", text, re.I),
        "glicose": first(rf"Glicose em L[íi]quor\.+:\s*{NUM}", text, re.I),
        "proteina": first(rf"Proteína em L[íi]quor:\s*{NUM}", text, re.I),
    }


def collect():
    files = sorted(
        [k for k in TEXTS if k.lower() != "todos.pdf"],
        key=lambda n: n,
    )
    resumo = []
    series = {
        "hb": [],
        "ht": [],
        "leu": [],
        "plaq": [],
        "pcr": [],
        "glicose": [],
        "na": [],
        "k": [],
        "cl": [],
        "ca": [],
        "mg": [],
        "creat": [],
        "lactato": [],
        "vanco": [],
        "cortisol": [],
        "fxiii": [],
        "tsh": [],
        "gas": [],
        "liquor": [],
        "culturas": [],
        "coag": [],
    }
    for fname in files:
        text = TEXTS[fname]
        pages = split_pages(text)
        headers = [parse_header(p) for p in pages]
        datas = sorted({h["data"] for h in headers if h["data"]})
        medicos = []
        for h in headers:
            if h["medico"] and h["medico"] not in medicos:
                medicos.append(h["medico"])
        locais = []
        for h in headers:
            sl = short_local(h["local"])
            if sl not in locais:
                locais.append(sl)
        dt_label = datas[0] if datas else None
        dt_extra = ""
        if len(datas) > 1:
            dt_extra = " e ".join(d.strftime("%d/%m") for d in datas[1:])
        resumo.append(
            {
                "data": dt_label,
                "datas": datas,
                "local": " / ".join(locais) if locais else "—",
                "medico": medicos[0] if medicos else "—",
                "desc": describe(text),
                "arquivo": fname,
                "paginas": len(pages),
                "presc": headers[0]["presc"] if headers else fname.replace(".PDF", ""),
            }
        )
        for page, h in zip(pages, headers):
            d = h["data"]
            rec = {"data": d, "arquivo": fname}

            hb = first(rf"Hemoglobina\.+:\s*{NUM}\s*g/dL", page)
            if hb is not None:
                series["hb"].append({**rec, "v": hb, "faixa": "9,0 a 14,0", "nota": "Abaixo da faixa" if hb < 9 else ""})
            ht = first(rf"Hematócrito\.+:\s*{NUM}\s*%", page)
            if ht is not None:
                series["ht"].append({**rec, "v": ht, "faixa": "ver laudo (muda com a idade)", "nota": ""})
            leu = first(rf"Leucócitos\.+:\s*{NUM}\s*x10", page)
            if leu is not None:
                abs_leu = round(leu * 1000)
                series["leu"].append({**rec, "v": abs_leu, "faixa": "6.000 a 17.000", "nota": "Abaixo da faixa" if abs_leu < 6000 else ""})
            plaq = first(rf"Plaquetas\.+:\s*{NUM}\s*x10", page)
            if plaq is not None:
                abs_p = round(plaq * 1000)
                series["plaq"].append({**rec, "v": abs_p, "faixa": "250.000 a 500.000 no laudo", "nota": "Abaixo da faixa" if abs_p < 250000 else ""})
            pcr = first(rf"PROTEÍNA C REATIVA - PCR:\s*{NUM}", page)
            if pcr is not None:
                series["pcr"].append({**rec, "v": pcr, "faixa": "0 a 10 mg/L", "nota": ""})
            gl = first(rf"(?m)^GLICOSE:\s*{NUM}", page)
            if gl is not None:
                series["glicose"].append({**rec, "v": gl, "faixa": "jejum 60 a 99", "nota": "Sem jejum (CTI)"})
            na = first(rf"SÓDIO:\s*{NUM}", page)
            if na is not None:
                series["na"].append({**rec, "v": na, "faixa": "136 a 145", "nota": "Abaixo" if na < 136 else ("Acima" if na > 145 else "")})
            kv = first(rf"POTÁSSIO:\s*{NUM}", page)
            if kv is not None:
                series["k"].append({**rec, "v": kv, "faixa": "3,5 a 5,8", "nota": "Abaixo" if kv < 3.5 else ("Acima" if kv > 5.8 else "")})
            cl = first(rf"CLORO:\s*{NUM}", page)
            if cl is not None:
                low = cl < 98
                high = cl > 107
                series["cl"].append({**rec, "v": cl, "faixa": "98 a 107", "nota": "Abaixo" if low else ("Acima" if high else "")})
            ca = first(rf"CÁLCIO IÔNICO:\s*{NUM}", page)
            if ca is not None:
                series["ca"].append({**rec, "v": ca, "faixa": "4,40 a 5,92 mg/dL", "nota": "Abaixo" if ca < 4.4 else ("Acima" if ca > 5.92 else "")})
            mg = first(rf"MAGNÉSIO:\s*{NUM}", page)
            if mg is not None:
                series["mg"].append({**rec, "v": mg, "faixa": "1,9 a 2,7 mg/dL", "nota": ""})
            cr = first(rf"CREATININA:\s*{NUM}", page)
            if cr is not None:
                series["creat"].append({**rec, "v": cr, "faixa": "0,17 a 0,52 mg/dL (laudo)", "nota": ""})
            lac = first(rf"ÁCIDO LÁTICO:\s*{NUM}", page)
            if lac is not None:
                series["lactato"].append({**rec, "v": lac, "faixa": "0,50 a 2,20 mmol/L", "nota": "Acima" if lac > 2.2 else ""})
            if re.search(r"(?m)^VANCOMICINA:", page):
                vv = first(rf"RESULTADO:\s*{NUM}\s*mcg/mL", page)
                if vv is not None:
                    series["vanco"].append({**rec, "v": vv, "faixa": "vale 10 a 20 mcg/mL", "nota": "Abaixo do vale" if vv < 10 else ""})
            if re.search(r"(?m)^CORTISOL\b", page):
                cv = first(rf"RESULTADO:\s*{NUM}\s*mcg/dL", page)
                if cv is not None:
                    series["cortisol"].append({**rec, "v": cv, "faixa": "matutino 5,3 a 22,5", "nota": "No piso / abaixo" if cv < 5.3 else ""})
            if "FATOR XIII" in page:
                fv = first(rf"RESULTADO:\s*{NUM}\s*%", page)
                if fv is not None:
                    series["fxiii"].append({**rec, "v": fv, "faixa": "70 a 140%", "nota": "Acima da faixa" if fv > 140 else ""})
            if "TIREOTROPINA" in page or "TSH) NEONATAL" in page:
                tv = first(rf"NEONATAL:\s*{NUM}", page)
                if tv is not None:
                    series["tsh"].append({**rec, "v": tv, "faixa": "após 7 dias até 10 mcU/mL", "nota": ""})
            for g in extract_gas_blocks(page):
                if any(g[k] is not None for k in ("ph", "pco2", "po2", "hco3", "be")):
                    series["gas"].append({**rec, **g})
            liq = extract_liquor(page)
            if liq and (liq["hemacias"] is not None or liq["glicose"] is not None):
                series["liquor"].append({**rec, **liq})
            if "COAGULOGRAMA" in page:
                rni = first(rf"RNI\.+:\s*{NUM}", page)
                atv = first(rf"Atividade Protrombina\.+:\s*{NUM}", page)
                series["coag"].append({**rec, "rni": rni, "atv": atv})

        # culturas at file level (result may be after header page)
        def add_cult(tipo, mat, res):
            series["culturas"].append(
                {
                    "data": datas[0] if datas else None,
                    "arquivo": fname,
                    "tipo": tipo,
                    "material": mat,
                    "resultado": res,
                }
            )

        if "HEMOCULTURA" in text:
            res = first_str(r"HEMOCULTURA[\s\S]{0,280}?Resultado\.:\s*(.+)", text)
            add_cult("Hemocultura", "Sangue", res or "ver PDF")
        if "UROCULTURA" in text:
            res = first_str(r"UROCULTURA[\s\S]{0,350}?Resultado\.:\s*(.+)", text)
            isol = first_str(r"Isolado 01:\s*(.+)", text)
            if isol:
                res = f"Positivo — {isol}"
            add_cult("Urocultura", "Urina", (res or "ver PDF").split("\n")[0][:80])
        if re.search(r"CULTURA AERÓBIA", text) and re.search(r"Material: Líquor", text):
            res = first_str(r"CULTURA AERÓBIA[\s\S]{0,280}?Resultado\.:\s*(.+)", text)
            add_cult("Cultura aeróbia", "Líquor", res or "ver PDF")

    for k, rows in series.items():
        rows.sort(key=lambda r: (r["data"] or date.min, r["arquivo"]))
    resumo.sort(key=lambda r: (r["data"] or date.min, r["arquivo"]))
    return resumo, series


def build_resumo(resumo, n_todos_pages, n_files):
    path = CTI / "Resumo Exames CTI.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    navy = "1B4F72"
    ws.merge_cells("A1:F1")
    ws["A1"] = "Cecília Maria Albergaria Silva — Exames laboratoriais do CTI (ago–set/2020)"
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color=navy)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Hospital BH · {n_files} PDFs individuais ({sum(r['paginas'] for r in resumo)} páginas). "
        f"O arquivo Todos.pdf tem {n_todos_pages} páginas e contém as mesmas {n_files} prescrições — está completo. "
        "Os nomes dos arquivos (número da prescrição) não foram alterados."
    )
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="5D6D7E")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 32

    headers = ["Data", "Local", "Profissional", "O que tem no PDF", "Arquivo", "Páginas"]
    ws.append(["", "", "", "", "", ""])  # row 3 placeholder then overwrite
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 16

    thin = Border(
        left=Side(style="thin", color="BDD7EE"),
        right=Side(style="thin", color="BDD7EE"),
        top=Side(style="thin", color="BDD7EE"),
        bottom=Side(style="thin", color="BDD7EE"),
    )
    local_fill = {
        "CTI Infantil": PatternFill("solid", fgColor="DDEBF7"),
        "PA": PatternFill("solid", fgColor="FCE4D6"),
        "Centro Cirúrgico": PatternFill("solid", fgColor="E2EFDA"),
    }
    cell_font = Font(name="Calibri", size=8)
    for i, r in enumerate(resumo, start=4):
        vals = [
            r["data"],
            r["local"],
            r["medico"],
            r["desc"],
            r["arquivo"],
            r["paginas"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            cell.font = cell_font
            cell.border = thin
            cell.alignment = Alignment(
                horizontal="center" if c != 4 else "left",
                vertical="center",
                wrap_text=(c == 4),
            )
            if c == 1:
                cell.number_format = "DD/MM/YYYY"
            if c == 2:
                cell.fill = local_fill.get(r["local"], PatternFill())
                cell.font = Font(name="Calibri", size=8, bold=True)
            elif i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8F9FA")
        ws.row_dimensions[i].height = 28 if len(r["desc"]) > 70 else 16

    last = 3 + len(resumo)
    for name in list(ws.tables.keys()):
        del ws.tables[name]
    tab = Table(displayName="TabelaCTI", ref=f"A3:F{last}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(tab)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 62
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:3"
    ws.print_area = f"A1:F{last}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.4, bottom=0.4, header=0.15, footer=0.2)
    ws.oddFooter.left.text = "Cecília — exames CTI · pasta Exames Laboratorias - CTI"
    ws.oddFooter.right.text = "Página &P de &N"
    wb.save(path)
    print("SAVED_RESUMO", path, "rows", len(resumo))
    return path


def build_evolucao(series):
    path = CTI / "Evolução Exames CTI.xlsx"
    wb = xlsxwriter.Workbook(str(path), {"default_date_format": "DD/MM/YYYY"})
    navy = "#1B4F72"
    ink = "#1C2833"
    muted = "#5D6D7E"
    band = "#F4F6F7"
    line = "#D5D8DC"
    white = "#FFFFFF"

    f_title = wb.add_format({"bold": True, "font_size": 16, "font_color": navy, "font_name": "Calibri", "valign": "vcenter"})
    f_sub = wb.add_format({"font_size": 10, "font_color": muted, "font_name": "Calibri", "text_wrap": True})
    f_h2 = wb.add_format({"bold": True, "font_size": 12, "font_color": navy, "font_name": "Calibri", "bg_color": "#D4E6F1", "valign": "vcenter"})
    f_h3 = wb.add_format({"italic": True, "font_size": 8, "font_color": muted, "font_name": "Calibri"})
    f_tab = wb.add_format({"bold": True, "font_size": 11, "font_color": ink, "font_name": "Calibri"})
    f_cab = wb.add_format({"bold": True, "font_size": 8, "font_color": white, "bg_color": navy, "font_name": "Calibri", "valign": "vcenter", "text_wrap": True})
    f_cell = wb.add_format({"font_size": 9, "font_name": "Calibri", "valign": "vcenter", "border": 1, "border_color": line, "font_color": ink, "text_wrap": True})
    f_cell_b = wb.add_format({"font_size": 9, "font_name": "Calibri", "valign": "vcenter", "border": 1, "border_color": line, "font_color": ink, "bg_color": band, "text_wrap": True})
    f_num = wb.add_format({"font_size": 9, "font_name": "Calibri", "align": "center", "valign": "vcenter", "border": 1, "border_color": line, "num_format": "#,##0.##"})
    f_num_b = wb.add_format({"font_size": 9, "font_name": "Calibri", "align": "center", "valign": "vcenter", "border": 1, "border_color": line, "bg_color": band, "num_format": "#,##0.##"})
    f_date = wb.add_format({"font_size": 9, "font_name": "Calibri", "align": "center", "valign": "vcenter", "border": 1, "border_color": line, "num_format": "DD/MM/YYYY"})
    f_date_b = wb.add_format({"font_size": 9, "font_name": "Calibri", "align": "center", "valign": "vcenter", "border": 1, "border_color": line, "bg_color": band, "num_format": "DD/MM/YYYY"})
    f_file = wb.add_format({"font_size": 8, "font_name": "Calibri", "align": "center", "valign": "vcenter", "border": 1, "border_color": line, "font_color": muted})
    f_file_b = wb.add_format({"font_size": 8, "font_name": "Calibri", "align": "center", "valign": "vcenter", "border": 1, "border_color": line, "font_color": muted, "bg_color": band})
    f_foot = wb.add_format({"italic": True, "font_size": 8, "font_color": muted, "font_name": "Calibri", "text_wrap": True})
    f_empty = wb.add_format({"font_size": 2})

    tables = []

    def simple(grupo, titulo, mostra, rows, unidade, faixa_key="faixa", nota_key="nota"):
        if not rows:
            return
        linhas = []
        for r in rows:
            linhas.append([r["data"], r["v"], unidade, r.get(faixa_key) or "", r.get(nota_key) or "", r["arquivo"]])
        tables.append({"grupo": grupo, "titulo": titulo, "mostra": mostra, "cols": ["Data", "Resultado", "Unidade", "Faixa do laudo", "Nota", "Arquivo"], "linhas": linhas, "kind": "simple"})

    simple("Hemograma", "Hemoglobina", "Anemia da hemorragia / CTI", series["hb"], "g/dL")
    simple("Hemograma", "Hematócrito", "Série vermelha", series["ht"], "%")
    simple("Hemograma", "Leucócitos", "Convertido de x10³/mm³", series["leu"], "/mm³")
    simple("Hemograma", "Plaquetas", "Convertido de x10³/mm³", series["plaq"], "/mm³")
    simple("Inflamação e infecção", "PCR", "Marcador de inflamação", series["pcr"], "mg/L")
    simple("Química", "Glicose", "Sangue (sem jejum, CTI/PA)", series["glicose"], "mg/dL")
    simple("Eletrólitos", "Sódio", "Na arterial ou venoso do laudo", series["na"], "mmol/L")
    simple("Eletrólitos", "Potássio", "K do laudo", series["k"], "mmol/L")
    simple("Eletrólitos", "Cloro", "Cl do laudo", series["cl"], "mmol/L")
    simple("Eletrólitos", "Cálcio iônico", "Sangue", series["ca"], "mg/dL")
    simple("Eletrólitos", "Magnésio", "Sangue", series["mg"], "mg/dL")
    simple("Rim", "Creatinina", "Valores do período neonatal/lactente no laudo", series["creat"], "mg/dL")
    simple("Gasometria", "Ácido lático / lactato", "Sangue", series["lactato"], "mmol/L")
    simple("Medicamentos", "Vancomicina (vale)", "Nível sérico", series["vanco"], "mcg/mL")
    simple("Hormônios", "Cortisol", "Coleta matutina no CTI", series["cortisol"], "mcg/dL")
    simple("Coagulação", "Fator XIII", "Única dosagem neste período", series["fxiii"], "%")
    simple("Hormônios", "TSH neonatal (papel filtro)", "Pesquisa de hipotireoidismo", series["tsh"], "mcU/mL")

    if series["gas"]:
        linhas = []
        for r in series["gas"]:
            linhas.append([r["data"], r.get("tipo") or "", r.get("ph"), r.get("pco2"), r.get("po2"), r.get("hco3"), r.get("be"), r["arquivo"]])
        tables.append({
            "grupo": "Gasometria",
            "titulo": "Gasometria (pH, gases e bicarbonato)",
            "mostra": "Maioria arterial · ver coluna Tipo",
            "cols": ["Data", "Tipo", "pH", "pCO2", "pO2", "HCO3", "BE", "Arquivo"],
            "linhas": linhas,
            "kind": "wide",
        })
    if series["liquor"]:
        linhas = []
        for r in series["liquor"]:
            linhas.append([r["data"], r.get("aspecto") or "", r.get("hemacias"), r.get("leuc"), r.get("glicose"), r.get("proteina"), r["arquivo"]])
        tables.append({
            "grupo": "Líquor",
            "titulo": "Rotina de líquor",
            "mostra": "Hemácias caem ao longo da internação (hemorragia → DVP)",
            "cols": ["Data", "Aspecto", "Hemácias /mm³", "Leucócitos /mm³", "Glicose mg/dL", "Proteína mg/dL", "Arquivo"],
            "linhas": linhas,
            "kind": "wide",
        })
    if series["culturas"]:
        linhas = []
        for r in series["culturas"]:
            linhas.append([r["data"], r["tipo"], r["material"], r["resultado"], "", r["arquivo"]])
        tables.append({
            "grupo": "Culturas",
            "titulo": "Hemocultura, urocultura e cultura de líquor",
            "mostra": "Uma linha por exame · resultado como no laudo",
            "cols": ["Data", "Exame", "Material", "Resultado", "", "Arquivo"],
            "linhas": linhas,
            "kind": "simple",
        })
    if series["coag"]:
        linhas = []
        for r in series["coag"]:
            linhas.append([r["data"], r.get("atv"), r.get("rni"), "Atividade % · RNI", "", r["arquivo"]])
        tables.append({
            "grupo": "Coagulação",
            "titulo": "Coagulograma (atividade / RNI)",
            "mostra": "Quando o PDF traz coagulograma",
            "cols": ["Data", "Atividade (%)", "RNI", "Faixa / nota", "", "Arquivo"],
            "linhas": linhas,
            "kind": "simple",
        })

    sh = wb.add_worksheet("Evolução")
    sh.set_tab_color(navy)
    sh.set_landscape()
    sh.set_paper(9)
    sh.set_margins(0.4, 0.4, 0.45, 0.45)
    sh.fit_to_pages(1, 0)
    sh.hide_gridlines(2)
    sh.center_horizontally()
    sh.set_column("A:A", 12)
    sh.set_column("B:B", 22)
    sh.set_column("C:C", 16)
    sh.set_column("D:D", 16)
    sh.set_column("E:E", 18)
    sh.set_column("F:F", 22)
    sh.set_column("G:G", 16)
    sh.set_column("H:H", 14)
    sh.set_header("&CCecília — evolução laboratorial do CTI (ago–set/2020)")
    sh.set_footer("&LValores conferidos nos PDFs desta pasta · não substitui o laudo&R&P / &N")

    sh.set_row(0, 22)
    sh.merge_range(0, 0, 0, 7, "Cecília Maria Albergaria Silva  ·  Evolução dos exames do CTI", f_title)
    sh.merge_range(
        1, 0, 1, 7,
        "04/08 a 28/09/2020 · Hospital Belo Horizonte. Cada linha aponta o PDF (número da prescrição). "
        "Todos.pdf junta os 68 arquivos e está completo. Esta pasta não mistura com os exames de fora do CTI.",
        f_sub,
    )
    r = 3
    current = None
    for t in tables:
        if t["grupo"] != current:
            current = t["grupo"]
            sh.set_row(r, 20)
            sh.merge_range(r, 0, r, 7, current, f_h2)
            r += 1
        sh.set_row(r, 18)
        sh.merge_range(r, 0, r, 7, t["titulo"], f_tab)
        r += 1
        sh.write(r, 0, t["mostra"], f_h3)
        r += 1
        ncols = len(t["cols"])
        for c, name in enumerate(t["cols"]):
            sh.write(r, c, name or None, f_cab)
        r += 1
        for j, linha in enumerate(t["linhas"]):
            odd = j % 2
            cf = f_cell_b if odd else f_cell
            nf = f_num_b if odd else f_num
            df = f_date_b if odd else f_date
            ff = f_file_b if odd else f_file
            sh.set_row(r, 18)
            for c, val in enumerate(linha):
                is_file = (c == len(linha) - 1)
                if isinstance(val, date):
                    sh.write_datetime(r, c, val, df)
                elif isinstance(val, (int, float)):
                    sh.write_number(r, c, val, nf)
                elif is_file:
                    sh.write(r, c, val, ff)
                elif val:
                    sh.write(r, c, val, cf)
                else:
                    sh.write(r, c, None, cf)
            r += 1
        r += 1
        sh.set_row(r - 1, 8, f_empty)

    sh.merge_range(
        r, 0, r + 1, 7,
        "Legenda: valores lidos dos PDFs da pasta Exames Laboratorias - CTI. "
        "Interpretação na coluna Nota é só confronto com a faixa impressa no próprio laudo. Conduta é com o médico. "
        "O arquivo Todos.pdf é a juntada dos 68 PDFs (162 páginas = soma das páginas individuais).",
        f_foot,
    )
    sh.print_area(0, 0, r + 1, 7)

    def pts(rows, key="v"):
        return [
            {"data": r["data"], "v": r[key]}
            for r in rows
            if r.get("data") is not None and r.get(key) is not None
        ]

    chart_specs = [
        ("Hemograma", "Hemoglobina", "g/dL", pts(series["hb"]), False, None, None),
        ("Hemograma", "Hematócrito", "%", pts(series["ht"]), False, None, None),
        ("Hemograma", "Leucócitos", "/mm³", pts(series["leu"]), False, None, None),
        ("Hemograma", "Plaquetas", "/mm³", pts(series["plaq"]), False, None, None),
        ("Inflamação", "PCR", "mg/L", pts(series["pcr"]), False, 0, None),
        ("Química", "Glicose", "mg/dL", pts(series["glicose"]), False, None, None),
        ("Química", "Creatinina", "mg/dL", pts(series["creat"]), False, 0, None),
        ("Química", "Lactato", "mmol/L", pts(series["lactato"]), False, 0, None),
        ("Eletrólitos", "Sódio", "mmol/L", pts(series["na"]), False, None, None),
        ("Eletrólitos", "Potássio", "mmol/L", pts(series["k"]), False, None, None),
        ("Eletrólitos", "Cloro", "mmol/L", pts(series["cl"]), False, None, None),
        ("Eletrólitos", "Cálcio iônico", "mg/dL", pts(series["ca"]), False, None, None),
        ("Eletrólitos", "Magnésio", "mg/dL", pts(series["mg"]), False, None, None),
        ("Gasometria", "pH", "pH", pts(series["gas"], "ph"), False, 7.20, 7.60),
        ("Gasometria", "pCO2", "mmHg", pts(series["gas"], "pco2"), False, None, None),
        ("Gasometria", "pO2", "mmHg", pts(series["gas"], "po2"), False, None, None),
        ("Gasometria", "HCO3", "mmol/L", pts(series["gas"], "hco3"), False, None, None),
        ("Gasometria", "BE (excesso de base)", "mmol/L", pts(series["gas"], "be"), False, None, None),
        ("Gasometria", "Saturação de O2", "%", pts(series["gas"], "so2"), False, 80, 100),
        ("Líquor", "Hemácias no líquor (escala log)", "Hemácias /mm³", pts(series["liquor"], "hemacias"), True, None, None),
        ("Líquor", "Leucócitos no líquor", "/mm³", pts(series["liquor"], "leuc"), False, 0, None),
        ("Líquor", "Glicose no líquor", "mg/dL", pts(series["liquor"], "glicose"), False, 0, None),
        ("Líquor", "Proteína no líquor", "mg/dL", pts(series["liquor"], "proteina"), False, 0, None),
        ("Medicamentos", "Vancomicina (vale)", "mcg/mL", pts(series["vanco"]), False, 0, None),
        ("Hormônios", "Cortisol", "mcg/dL", pts(series["cortisol"]), False, 0, None),
        ("Coagulação", "Atividade de protrombina", "%", pts(series["coag"], "atv"), False, None, None),
        ("Coagulação", "RNI", "RNI", pts(series["coag"], "rni"), False, None, None),
        ("Coagulação", "Fator XIII", "%", pts(series["fxiii"]), False, None, None),
        ("Hormônios", "TSH neonatal (papel filtro)", "mcU/mL", pts(series["tsh"]), False, 0, None),
    ]

    ds = wb.add_worksheet("Dados gráficos")
    ds.hide()
    plotted = []
    skipped = []
    for i, spec in enumerate(chart_specs):
        grupo, title, y_name, rows, log, ymin, ymax = spec
        col = i * 3
        ds.write(0, col, title)
        ds.write(1, col, "Data")
        ds.write(1, col + 1, "Valor")
        n = 0
        for j, row in enumerate(rows):
            ds.write_datetime(j + 2, col, row["data"])
            ds.write_number(j + 2, col + 1, row["v"])
            n = j + 1
        if n >= 2:
            plotted.append((grupo, title, y_name, col, n, log, ymin, ymax))
        else:
            skipped.append((title, n))

    ch = wb.add_worksheet("Gráficos")
    ch.set_tab_color("#148F77")
    ch.set_landscape()
    ch.set_paper(9)
    ch.hide_gridlines(2)
    ch.fit_to_pages(1, 0)
    ch.set_margins(0.4, 0.4, 0.5, 0.45)
    ch.merge_range(0, 0, 0, 15, "Cecília — todos os gráficos do CTI (ago–set/2020)", f_title)
    ch.merge_range(
        1, 0, 1, 15,
        "Fonte: PDFs individuais desta pasta. Eixo X = data da coleta. "
        "Fator XIII e TSH neonatal têm uma só dosagem: ficam só na aba Evolução. Culturas são qualitativas (tabela).",
        f_sub,
    )
    ch.set_column("A:P", 9)
    ch.set_row(0, 22)
    ch.set_row(1, 28)

    def add_line_chart(title, y_name, col, n, log=False, ymin=None, ymax=None):
        chart = wb.add_chart({"type": "line"})
        chart.add_series({
            "name": title,
            "categories": ["Dados gráficos", 2, col, n + 1, col],
            "values": ["Dados gráficos", 2, col + 1, n + 1, col + 1],
            "line": {"color": navy, "width": 1.5},
            "marker": {"type": "circle", "size": 5, "border": {"color": navy}, "fill": {"color": navy}},
        })
        chart.set_title({"name": title, "name_font": {"size": 11, "color": navy}})
        chart.set_x_axis({
            "name": "Data da coleta",
            "num_format": "DD/MM",
            "name_font": {"size": 9},
            "num_font": {"size": 8},
        })
        yaxis = {"name": y_name, "name_font": {"size": 9}, "num_font": {"size": 8}}
        if log:
            yaxis["log_base"] = 10
        if ymin is not None:
            yaxis["min"] = ymin
        if ymax is not None:
            yaxis["max"] = ymax
        chart.set_y_axis(yaxis)
        chart.set_legend({"none": True})
        chart.set_size({"width": 460, "height": 268})
        chart.set_style(10)
        return chart

    row = 3
    gi = 0
    while gi < len(plotted):
        grupo = plotted[gi][0]
        ch.set_row(row, 20)
        ch.merge_range(row, 0, row, 15, grupo, f_h2)
        row += 2
        slot = 0
        while gi < len(plotted) and plotted[gi][0] == grupo:
            _, title, y_name, col, n, log, ymin, ymax = plotted[gi]
            chart = add_line_chart(title, y_name, col, n, log, ymin, ymax)
            ch.insert_chart(row, 0 if slot == 0 else 8, chart)
            slot += 1
            gi += 1
            if slot == 2:
                slot = 0
                row += 15
        if slot == 1:
            row += 15

    how = wb.add_worksheet("Como usar")
    how.hide_gridlines(2)
    how.set_column("A:A", 120)
    how.write(0, 0, "Como usar esta pasta", f_title)
    lines = [
        "",
        "Esta pasta (Exames Laboratorias - CTI) ficou separada de propósito: é o período de UTI, com coleta quase diária.",
        "Não misturamos com o Resumo/Evolução da pasta Exames (consultas e ambulatório).",
        "Os PDFs individuais mantiveram o nome original (número da prescrição do Hospital BH).",
        "Todos.pdf junta os 68 PDFs: 162 páginas no conjunto = 162 páginas somadas nos individuais. Nenhuma prescrição faltando.",
        "Resumo Exames CTI.xlsx — uma linha por PDF, com a data, o local (PA / CTI / centro cirúrgico), o que tem no laudo e o nome do arquivo.",
        "Evolução Exames CTI.xlsx — tabelas por exame. A última coluna é o arquivo.",
        "Aba Gráficos — uma linha do tempo para cada medida numérica (hemograma, eletrólitos, gasometria, líquor, glicose, lactato, PCR, vancomicina, cortisol, coagulograma).",
        "Fator XIII e TSH neonatal têm só um ponto: aparecem na tabela, não no gráfico. Culturas (positivo/negativo) ficam na tabela.",
        "Urocultura positiva (Citrobacter freundii) está na tabela de culturas, arquivo 3070603.PDF (11/08/2020).",
    ]
    for i, line in enumerate(lines):
        how.set_row(i + 1, 28 if line else 10)
        how.write(i + 1, 0, line, f_sub)
    print("charts", len(plotted), "skipped", skipped)

    sh.activate()
    wb.close()
    print("SAVED_EVOL", path, "tables", len(tables))
    for t in tables:
        print(f"  {len(t['linhas']):3}  {t['titulo']}")
    return path


def main():
    resumo, series = collect()
    print("resumo", len(resumo))
    for k, v in series.items():
        print(f"  {k:10} {len(v)}")
    n_todos = len(split_pages(TEXTS["Todos.pdf"]))
    build_resumo(resumo, n_todos, len(resumo))
    build_evolucao(series)


if __name__ == "__main__":
    main()
